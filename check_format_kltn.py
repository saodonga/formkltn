#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_format_kltn.py — Kiểm tra định dạng Khóa Luận Tốt Nghiệp (KLTN)
========================================================================
Dựa theo "Hướng dẫn trình bày ĐATN/KLTN" của Trường Đại học Thủy Lợi

Kiểm tra các yêu cầu:
  1. Khổ giấy A4 (21 x 29.7 cm)
  2. Lề: Trái 3cm, Phải 2cm, Trên 2.5cm, Dưới 2.5cm
  3. Font chữ Times New Roman cho toàn bộ văn bản
  4. Cỡ chữ:
     - Heading 1 (Chương): cỡ 14pt, đậm, in hoa
     - Heading 2: cỡ 13pt, đậm
     - Heading 3: cỡ 13pt, đậm + nghiêng
     - Heading 4: cỡ 13pt, nghiêng
     - Nội dung (Content/Body): cỡ 13pt
     - Caption: cỡ 12pt, nghiêng
     - Tối thiểu 10pt, tối đa 13pt cho nội dung thường
  5. Giãn dòng:
     - Nội dung: 1.5 lines
     - Heading: single
     - Caption: 1.3 multiple
  6. Canh lề nội dung: hai bên (justify)
  7. Cấu trúc bắt buộc: Trang bìa, Lời cam đoan, Mục lục,
     Danh mục hình, Danh mục bảng, Nội dung chương, Tài liệu tham khảo
  8. Đánh số trang đúng quy định (i,ii,iii... phần mở đầu; 1,2,3... phần nội dung)

Cách dùng:
  python check_format_kltn.py                        # Chọn file qua hộp thoại
  python check_format_kltn.py "path/to/kltn.docx"   # Truyền đường dẫn trực tiếp
  python check_format_kltn.py "folder/"             # Quét toàn bộ thư mục
"""

import os
import re
import sys
import json
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, field
from typing import Optional

# ─── Tự động cài thư viện nếu thiếu ─────────────────────────────
def _ensure_pkg(pkg_name, import_name=None):
    import_name = import_name or pkg_name
    try:
        __import__(import_name)
    except ImportError:
        print(f"[!] Thiếu '{pkg_name}'. Đang cài...")
        os.system(f"{sys.executable} -m pip install {pkg_name} -q")

_ensure_pkg("python-docx", "docx")
_ensure_pkg("openpyxl")

import docx
from docx.shared import Pt, Cm, Emu
from docx.enum.text import WD_LINE_SPACING, WD_ALIGN_PARAGRAPH
from openpyxl import Workbook
from openpyxl.styles import Font as XFont, PatternFill, Alignment, Border, Side

import zipfile
# ── Patch: Bỏ qua lỗi Bad CRC-32 của thư viện zipfile khi đọc file Word có chứa ảnh bị lỗi ngầm
_original_zip_ext_init = zipfile.ZipExtFile.__init__
def _patched_zip_ext_init(self, *args, **kwargs):
    _original_zip_ext_init(self, *args, **kwargs)
    self._expected_crc = None  # Tắt kiểm tra CRC để ngăn python-docx crash
zipfile.ZipExtFile.__init__ = _patched_zip_ext_init

from openpyxl.utils import get_column_letter


# ════════════════════════════════════════════════════════════════
#  CẤU HÌNH CHUẨN (từ hướng dẫn)
# ════════════════════════════════════════════════════════════════
class Std:
    """Các hằng số chuẩn theo hướng dẫn trình bày KLTN."""
    PAPER_W_CM   = 21.0
    PAPER_H_CM   = 29.7
    MARGIN_L_CM  = 3.0
    MARGIN_R_CM  = 2.0
    MARGIN_T_CM  = 2.5
    MARGIN_B_CM  = 2.5
    TOLERANCE_CM = 0.55   # dung sai 5.5mm (chấp nhận sai số nhỏ do Word rounding)

    FONT_MAIN    = "Times New Roman"

    H1_SIZE_PT   = 14
    H2_SIZE_PT   = 13
    H3_SIZE_PT   = 13
    H4_SIZE_PT   = 13
    BODY_SIZE_PT = 13
    CAPTION_SIZE_PT = 12
    BODY_LINE_SPACING = 1.5
    H1_SPACE_BEFORE_PT = 24
    H1_SPACE_AFTER_PT  = 24
    H2_SPACE_BEFORE_PT = 6
    H2_SPACE_AFTER_PT  = 12
    H3_SPACE_BEFORE_PT = 6
    H3_SPACE_AFTER_PT  = 12
    BODY_SPACE_BEFORE_PT = 10
    BODY_SPACE_AFTER_PT  = 0

    REQUIRED_SECTIONS = [
        "lời cam đoan",
        "mục lục",
        "danh mục",
        "tài liệu tham khảo",
    ]

    REQUIRED_HEADING_KEYWORDS = [
        "tổng quan",        # chương 1: tổng quan / giới thiệu
        "thực trạng",       # chương 2
        "giải pháp",        # chương 3 / chương kết quả
    ]


def _emu_to_cm(emu) -> Optional[float]:
    if emu is None:
        return None
    return emu / 914400 * 2.54


def _pt_from_emu(emu) -> Optional[float]:
    """Chuyển EMU (twentieth of a point × 12700) sang pt."""
    if emu is None:
        return None
    return emu / 12700


def _pt_to_emu(pt: float) -> int:
    return int(pt * 12700)


# ════════════════════════════════════════════════════════════════
#  DATA CLASSES
# ════════════════════════════════════════════════════════════════
@dataclass
class Issue:
    category: str       # Nhóm lỗi
    severity: str       # "ERROR" | "WARNING" | "INFO"
    message: str        # Mô tả
    location: str = ""  # Vị trí (tên style, số đoạn…)
    suggestion: str = ""


@dataclass
class CheckResult:
    filepath: str
    student_name: str = ""
    student_id: str = ""
    title: str = ""
    advisor: str = ""
    issues: list = field(default_factory=list)
    pass_count: int = 0
    warn_count: int = 0
    error_count: int = 0
    score: int = 0          # /100

    @property
    def letter_grade(self) -> str:
        s = self.score
        if s >= 90: return "A+"
        if s >= 85: return "A"
        if s >= 80: return "B+"
        if s >= 70: return "B"
        if s >= 65: return "C+"
        if s >= 55: return "C"
        if s >= 50: return "D+"
        if s >= 40: return "D"
        if s >= 20: return "E+"
        return "E"

    def summary(self) -> str:
        e = sum(1 for i in self.issues if i.severity == "ERROR")
        w = sum(1 for i in self.issues if i.severity == "WARNING")
        return f"Lỗi: {e} | Cảnh báo: {w}"


# ════════════════════════════════════════════════════════════════
#  CẤU HÌNH (config_kltn.json)
# ════════════════════════════════════════════════════════════════
def _load_config() -> dict:
    """Load config_kltn.json nếu tồn tại bên cạnh script."""
    config_path = Path(__file__).parent / "config_kltn.json"
    if config_path.exists():
        try:
            with open(config_path, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def _normalize_advisor(name: str) -> str:
    """Chuẩn hóa tên GVHD: bỏ học vị, số thứ tự, khoảng trắng thừa, chữ thường."""
    # Bỏ số thứ tự đầu: "1.", "2.", "1)", ...
    name = re.sub(r'^\s*\d+[.)]\s*', '', name).strip()
    # Bỏ học vị phổ biến
    name = re.sub(
        r'\b(PGS\.?TS\.?|GS\.?TS\.?|TS\.?|ThS\.?|CN\.?|NCS\.?)\s*',
        '', name, flags=re.IGNORECASE
    ).strip()
    # Bỏ khoảng trắng thừa, chuỗi placeholder
    name = re.sub(r'\s+', ' ', name).strip()
    return name.lower()

def _advisor_in_list(raw_name: str, advisor_list: list) -> bool:
    """Kiểm tra tên GVHD (đã bình thường hóa) có trong danh sách không."""
    norm_raw = _normalize_advisor(raw_name)
    if len(norm_raw) < 3:
        return False
    for ref in advisor_list:
        norm_ref = _normalize_advisor(ref)
        # So khớp: tên ref nằm trong tên raw hoặc ngược lại
        if norm_ref and (norm_ref in norm_raw or norm_raw in norm_ref):
            return True
    return False


# ════════════════════════════════════════════════════════════════
#  CÔNG CỤ KIỂM TRA
# ════════════════════════════════════════════════════════════════
class KLTNChecker:
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.doc = docx.Document(filepath)
        self.result = CheckResult(filepath=filepath)
        self._paras = self.doc.paragraphs
        self._all_text_paras = [p for p in self._paras if p.text.strip()]

    def check_all(self) -> CheckResult:
        """Chạy toàn bộ các kiểm tra."""
        self._extract_cover_info()
        self._check_page_setup()
        self._check_cover_pages()   # <-- kiểm tra trang bìa
        self._check_font_and_styles()
        self._check_structure()
        self._check_headings()
        self._check_body_text()
        self._check_captions()
        self._check_references()
        self._check_abbreviations_and_quotes()
        self._check_ai_copy_anomalies()
        self._compute_score()
        return self.result

    # ── 0. Trích thông tin trang bìa ─────────────────────────────
    def _extract_cover_info(self):
        lines = [p.text.strip() for p in self._paras[:80] if p.text.strip()]

        # Tên đề tài
        title_idx = None
        STOP = re.compile(r'giảng viên|gvhd|sinh viên|mssv|lớp|khóa|hà nội|tp\.',
                          re.IGNORECASE)
        ANCHOR = re.compile(r'khóa luận|đề tài|luận văn|báo cáo|graduation',
                            re.IGNORECASE)
        for i, line in enumerate(lines):
            if ANCHOR.search(line) and not title_idx:
                title_idx = i
        if title_idx:
            title_lines = []
            for line in lines[title_idx+1:title_idx+6]:
                if STOP.search(line):
                    break
                if len(line) >= 5:
                    title_lines.append(line)
            self.result.title = ' '.join(title_lines)[:250]

        # MSSV
        mssv_m = re.search(r'(?:MSSV|mã số)[:\s]+(\d{8,12})',
                            '\n'.join(lines[:60]), re.IGNORECASE)
        if mssv_m:
            self.result.student_id = mssv_m.group(1)
        else:
            m = re.search(r'\b(\d{10})\b', '\n'.join(lines[:60]))
            if m:
                self.result.student_id = m.group(1)

        # Họ tên
        name_m = re.search(r'(?:sinh viên|họ và tên|họ tên|tên sinh viên)[:\s]+([^\n]{5,60})',
                            '\n'.join(lines[:60]), re.IGNORECASE)
        if name_m:
            name = name_m.group(1).strip()
            name = re.split(r'\s{2,}|\t|\d{5,}', name)[0].strip()
            self.result.student_name = name[:60]

        # GVHD
        gvhd_m = re.search(
            r'(?:giảng viên hướng dẫn|gvhd|người hướng dẫn|hướng dẫn khoa học)[:\s.]+([^\n]{5,80})',
            '\n'.join(lines[:60]), re.IGNORECASE)
        if gvhd_m:
            gvhd = gvhd_m.group(1).strip()
            gvhd = re.sub(r'^(pgs\.?ts\.?|gs\.?ts\.?|ts\.?|ths\.?|gs\.?)\s+', '', gvhd,
                          flags=re.IGNORECASE).strip()
            self.result.advisor = gvhd[:60]

    # ── 0b. Kiểm tra định dạng trang bìa ────────────────────────
    def _check_cover_pages(self):
        """
        Kiểm tra định dạng trang bìa chính (trang 1) và trang bìa phụ (trang 2).

        Theo file mẫu chuẩn 2026_Mau_viet_chuan_KLTN.docx:
          • Tên đề tài  : Times New Roman, 16pt, Bold, IN HOA, canh giữa, màu xanh #0000FF
          • Họ và tên   : Times New Roman, 14pt, không đậm, IN HOA, canh giữa, màu xanh
          • KHÓA LUẬN TỐT NGHIỆP : 14pt, IN HOA, không đậm
        """
        # ── Xác định giới hạn trang bìa (trước 'lời cam đoan') ──
        cover_end = 80  # tối đa
        COVER_END_MARKERS = re.compile(
            r'lời cam đoan|mục lục|danh mục|chương 1|chương i\b',
            re.IGNORECASE
        )
        for i, p in enumerate(self._paras[:100]):
            if COVER_END_MARKERS.search(p.text):
                cover_end = i
                break

        cover_paras = self._paras[:cover_end]

        # ── Kiểm tra tên Bộ ───────────────────────────────────────
        MINISTRY_INCORRECT = re.compile(
            r'bộ\s+nông\s+nghiệp\s+và\s+ptnt|bộ\s+nông\s+nghiệp\s+&\s+ptnt|bộ\s+nn\s*&\s*ptnt',
            re.IGNORECASE
        )
        for pi, para in enumerate(cover_paras):
            text = para.text.strip()
            if not text:
                continue
            if MINISTRY_INCORRECT.search(text):
                self.result.issues.append(Issue(
                    "Trang bìa — Tên CƠ QUAN", "WARNING",
                    f"Ghi sai cơ quan chủ quản: '{text[:50]}...'",
                    "Trang bìa",
                    "Yêu cầu chuyển chữ 'Bộ nông nghiệp và ptnt' thành 'BỘ NÔNG NGHIỆP VÀ MÔI TRƯỜNG' theo quy định mới."
                ))

        # ── Helper: lấy thuộc tính run đầu tiên có text ──────────
        def _run_props(para):
            """Trả về (font_size_pt, is_bold, color_hex, is_all_caps) của run đầu."""
            for run in para.runs:
                if not run.text.strip():
                    continue
                sz   = run.font.size
                sz_pt = round(sz / 12700, 1) if sz else None
                bold  = run.font.bold
                caps  = run.font.all_caps
                try:
                    color = str(run.font.color.rgb) if run.font.color and run.font.color.type else None
                except Exception:
                    color = None
                return sz_pt, bold, color, caps
            return None, None, None, None

        def _is_all_upper(text: str) -> bool:
            """Kiểm tra toàn bộ ký tự chữ cái có viết hoa không."""
            alphas = [c for c in text if c.isalpha()]
            return bool(alphas) and all(c.isupper() for c in alphas)



        # ── Tìm tên đề tài trong trang bìa ──────────────────────
        # Đặc điểm: paragraph có font size lớn nhất (>=15pt) trong trang bìa
        # KHÔNG phải tên trường, không phải nhãn cố định
        EXCLUDE_PATTERNS = re.compile(
            r'^(trường|bộ giáo|đại học thủy lợi|khoa |khóa luận tốt nghiệp'
            r'|báo cáo thực tập|hà nội|tp\.|ngành|mã số|người hướng|'
            r'sinh viên|họ và tên|mssv|gvhd|năm 20|\d{4}$)',
            re.IGNORECASE
        )

        title_candidates = []
        title_candidates_small = []   # đề tài có font nhỏ vẫn đưa vào check
        for pi, para in enumerate(cover_paras):
            text = para.text.strip()
            if not text or len(text) < 5:
                continue
            if EXCLUDE_PATTERNS.match(text):
                continue
            sz_pt, bold, color, caps = _run_props(para)
            # Đề tài chuẩn là 16pt; ít nhất phải >= 15pt để không nhầm với tên SV (14pt)
            if sz_pt and sz_pt >= 15:
                title_candidates.append((pi, para, text, sz_pt, bold, color, caps))
            elif sz_pt and sz_pt >= 12 and bold:
                # Font nhỏ nhưng bold → có thể là đề tài bị định dạng sai
                title_candidates_small.append((pi, para, text, sz_pt, bold, color, caps))

        # Nếu không tìm thấy đề tài lớn, thử tìm trong nhóm nhỏ
        if not title_candidates:
            if title_candidates_small:
                # Dùng đề tài font nhỏ nhất (có thể sẽ bị lỗi cỡ chữ)
                title_candidates = title_candidates_small
                self.result.issues.append(Issue(
                    "Trang bìa — Tên đề tài", "WARNING",
                    "Có vẻ trang bìa không có đoạn nào có cỡ chữ ≥14pt. "
                    "Tên đề tài thường phải là 16pt, Bold, màu xanh, IN HOA.",
                    "Trang bìa",
                    "Đặt cỡ chữ 16pt cho tên đề tài trên trang bìa."
                ))
            else:
                self.result.issues.append(Issue(
                    "Trang bìa", "WARNING",
                    "Không tìm thấy tên đề tài trên trang bìa (paragraph cỡ chữ ≥14pt).",
                    "Trang bìa",
                    "Đảm bảo tên đề tài được định dạng 16pt, Bold, màu xanh, IN HOA trên trang bìa."
                ))
                return

        # Ưu tiên paragraph có font lớn nhất (đó là tên đề tài)
        title_candidates.sort(key=lambda x: (-(x[3] or 0), x[0]))
        # Lấy tối đa 2 (bìa chính + bìa phụ)
        checked_texts = set()
        found_titles  = []
        for candidate in title_candidates:
            pi, para, text, sz_pt, bold, color, caps = candidate
            if text not in checked_texts:
                found_titles.append(candidate)
                checked_texts.add(text)
            if len(found_titles) >= 2:
                break

        # ── Kiểm tra từng tên đề tài tìm được ───────────────────
        STD_TITLE_SIZE  = 16.0   # chuẩn theo file mẫu
        STD_TITLE_BOLD  = True
        STD_TITLE_BLUE  = True
        STD_TITLE_UPPER = True
        SIZE_TOL = 1.0           # dung sai ±1pt

        for idx, (pi, para, text, sz_pt, bold, color, caps) in enumerate(found_titles):
            page_label = f"Trang bìa {'chính' if idx == 0 else 'phụ'} (đoạn {pi})"
            short_text = text[:60] + ('…' if len(text) > 60 else '')

            # 1. Cỡ chữ
            if sz_pt is None:
                self.result.issues.append(Issue(
                    "Trang bìa — Tên đề tài", "WARNING",
                    f"Không xác định được cỡ chữ tên đề tài: '{short_text}'",
                    page_label,
                    f"Đặt cỡ chữ {STD_TITLE_SIZE:.0f}pt cho tên đề tài."
                ))
            elif abs(sz_pt - STD_TITLE_SIZE) > SIZE_TOL:
                self.result.issues.append(Issue(
                    "Trang bìa — Tên đề tài", "ERROR",
                    f"Cỡ chữ tên đề tài {sz_pt}pt ≠ chuẩn {STD_TITLE_SIZE:.0f}pt: '{short_text}'",
                    page_label,
                    f"Chọn tên đề tài → đặt cỡ chữ {STD_TITLE_SIZE:.0f}pt (Times New Roman)."
                ))

            # 2. In đậm
            if bold is False:  # False = tường minh không đậm; None = kế thừa
                self.result.issues.append(Issue(
                    "Trang bìa — Tên đề tài", "ERROR",
                    f"Tên đề tài không in đậm (Bold): '{short_text}'",
                    page_label,
                    "Bôi đen tên đề tài trên trang bìa → Ctrl+B."
                ))

            # 3. IN HOA (all caps)
            is_upper = _is_all_upper(text) or bool(caps)
            if not is_upper:
                self.result.issues.append(Issue(
                    "Trang bìa — Tên đề tài", "ERROR",
                    f"Tên đề tài không IN HOA: '{short_text}'",
                    page_label,
                    "Viết tên đề tài bằng chữ IN HOA toàn bộ trên trang bìa. "
                    "Có thể dùng Format → Font → All Caps, hoặc gõ trực tiếp chữ hoa."
                ))



        # ── Kiểm tra Họ và Tên trên trang bìa ───────────────────
        HO_TEN_PATTERNS = re.compile(
            r'^h[oọ]\s+v[aà]\s+t[eê]n|^t[eê]n\s+sinh\s+vi[eê]n',
            re.IGNORECASE
        )
        for para in cover_paras:
            text = para.text.strip()
            if not text or not HO_TEN_PATTERNS.match(text):
                continue
            sz_pt, bold, color, caps = _run_props(para)
            short = text[:40]

            # Họ tên: cỡ 14pt, IN HOA, màu xanh
            if sz_pt and abs(sz_pt - 14.0) > 1.5:
                self.result.issues.append(Issue(
                    "Trang bìa — Họ tên", "WARNING",
                    f"Cỡ chữ 'Họ và tên' {sz_pt}pt ≠ chuẩn 14pt: '{short}'",
                    "Trang bìa",
                    "Đặt cỡ chữ 14pt cho trường 'Họ và Tên' trên trang bìa."
                ))

            if not _is_all_upper(text) and not caps:
                self.result.issues.append(Issue(
                    "Trang bìa — Họ tên", "WARNING",
                    f"Trường 'Họ và tên' không IN HOA: '{short}'",
                    "Trang bìa",
                    "Viết tên IN HOA toàn bộ trên trang bìa cho trường họ và tên."
                ))
        # ── Kiểm tra Họ và Tên sinh viên ─────────────────────────
        # Placeholder patterns khi SV chưa điền
        NAME_PLACEHOLDER = re.compile(
            r'^h[oọ]\s+v[aà]\s+t[eê]n\s*$'          # "họ và tên"
            r'|^h[oọ]\s+t[eê]n\s+sinh\s+vi[eê]n\s*$' # "họ tên sinh viên"
            r'|^student\s+name\s*$'                    # "student name"
            r'|^\.{3,}\s*$',                           # "..."
            re.IGNORECASE
        )
        # Tìm paragraph tên SV: 14pt, trước đoạn 16pt
        # Dùng EXCLUDE riêng (KHÔNG loại "họ và tên" để detect placeholder)
        NAME_SECTION_EXCLUDE = re.compile(
            r'^(trường|bộ giáo|đại học thủy lợi|khoa\s|khóa luận tốt nghiệp'
            r'|báo cáo thực tập|hà nội|tp\.|ngành|mã số|người hướng'
            r'|mssv|gvhd|năm 20|\d{4}$)',
            re.IGNORECASE
        )
        first_title_idx = found_titles[0][0] if found_titles else cover_end
        name_para_found = None
        for pi, para in enumerate(cover_paras):
            if pi >= first_title_idx:
                break
            text = para.text.strip()
            if not text or len(text) < 2:
                continue
            if NAME_SECTION_EXCLUDE.match(text):
                continue
            sz_pt, bold, color, caps = _run_props(para)
            if sz_pt and abs(sz_pt - 14.0) <= 1.5:
                name_para_found = (pi, text, sz_pt)
                break  # lấy đoạn 14pt đầu tiên trước đề tài

        if name_para_found:
            _, name_text, _ = name_para_found
            if NAME_PLACEHOLDER.match(name_text):
                self.result.issues.append(Issue(
                    "Trang bìa — Họ tên SV", "ERROR",
                    f"Chưa điền họ và tên sinh viên (vẫn là placeholder: '{name_text}').",
                    "Trang bìa",
                    "Thay text 'HỌ VÀ TÊN' bằng họ tên thực của sinh viên (dưới 35 ký tự, IN HOA)."
                ))
            elif len(name_text) > 35:
                self.result.issues.append(Issue(
                    "Trang bìa — Họ tên SV", "WARNING",
                    f"Họ và tên sinh viên quá dài ({len(name_text)} ký tự > 35): '{name_text}'",
                    "Trang bìa",
                    "Họ tên sinh viên thường dưới 35 ký tự. Kiểm tra có bị thừa thông tin không."
                ))

        # ── Kiểm tra Tên đề tài ───────────────────────────────────
        TITLE_PLACEHOLDER = re.compile(
            r'^(nhập\s+)?tên\s+đề\s+tài'   # "tên đề tài ..." / "nhập tên đề tài ..."
            r'|^(nhập\s+)?title'
            r'|tên\s+đề\s+tài\s+kltn',
            re.IGNORECASE
        )
        TITLE_MIN_LEN = _load_config().get("_title_min_length", 50)

        for idx, (pi, para, text, sz_pt, bold, color, caps) in enumerate(found_titles):
            short = text[:70] + ("…" if len(text) > 70 else "")
            page_label = f"Trang bìa {'chính' if idx == 0 else 'phụ'} (đoạn {pi})"

            # Kiểm tra còn là placeholder
            if TITLE_PLACEHOLDER.search(text):
                self.result.issues.append(Issue(
                    "Trang bìa — Tên đề tài", "ERROR",
                    f"Chưa điền tên đề tài (vẫn còn cụm từ: '{short}').",
                    page_label,
                    "Bỏ cụm từ 'TÊN ĐỀ TÀI KLTN' và thay thế bằng tên Đề tài KLTN thực tế vào (IN HOA, in đậm, 16pt)."
                ))
                continue  # bỏ qua các check khác với placeholder

            # Kiểm tra độ dài
            if len(text) < TITLE_MIN_LEN:
                self.result.issues.append(Issue(
                    "Trang bìa — Tên đề tài", "WARNING",
                    f"Tên đề tài quá ngắn ({len(text)} ký tự < {TITLE_MIN_LEN}): '{short}'",
                    page_label,
                    "Tên đề tài KLTN thường từ 50 ký tự trở lên. Kiểm tra có bị cắt ngắn không."
                ))

            # Kiểm tra dấu ngoặc bao bọc
            has_quotes = bool(re.search(r'^["\'\u201c\u2018\u00ab]|["\'\u201d\u2019\u00bb]$', text.strip()))
            if has_quotes:
                self.result.issues.append(Issue(
                    "Trang bìa — Tên đề tài", "WARNING",
                    f"Tên đề tài không được đặt trong dấu ngoặc: '{short}'",
                    page_label,
                    "Xóa dấu ngoặc kép / ngoặc đơn bao bọc tên đề tài trên trang bìa."
                ))

        # ── Kiểm tra Người hướng dẫn (GVHD) ─────────────────────
        config = _load_config()
        advisor_list = [a.strip() for a in config.get("advisors", []) if a.strip()]

        # Tìm đoạn "NGƯỜI HƯỚNG DẪN" trong trang bìa (ưu tiên trang bìa phụ)
        GVHD_PAT = re.compile(
            r'hướng\s+dẫn|gvhd',
            re.IGNORECASE
        )
        gvhd_lines = []   # list tên GVHD tìm được
        gvhd_found_para = False

        for pi, para in enumerate(cover_paras):
            text = para.text.strip()
            if not text:
                continue

            if GVHD_PAT.search(text):
                gvhd_found_para = True
                # Trích tên từ cùng đoạn: phần sau dấu ":" hoặc "1."
                rest = re.split(r'[:\t]\s*', text, maxsplit=1)[-1].strip()
                if rest and len(rest) > 3 and not GVHD_PAT.search(rest):
                    # Có thể là "1. TS. Nguyễn..." hoặc nhiều tên trên 1 dòng
                    parts = re.split(r'\s{3,}|\t|(?<=\))\s+(?=\d+\.)', rest)
                    for part in parts:
                        part = part.strip()
                        if len(part) > 3:
                            gvhd_lines.append(part)
                # Tìm tiếp vài đoạn ngay sau (cho cấu trúc "1.\n2.")
                for j in range(pi + 1, min(pi + 6, len(cover_paras))):
                    next_text = cover_paras[j].text.strip()
                    if not next_text:
                        continue
                    # Dừng nếu gặp phần khác (Ngành, Mã số, Hà Nội, ...)
                    if re.match(r'^(ng[aà]nh|m[aã]\s*s[oố]|h[aà]\s+n[oộ]i|n[aă]m\s+20)', next_text, re.IGNORECASE):
                        break
                    # Dòng tiếp là tên GVHD thêm (dạng "2. ThS. Trần...")
                    if re.match(r'^\d+[.)]\s*', next_text) or re.match(r'^(ts|ths|pgs|gs)\b', next_text, re.IGNORECASE):
                        gvhd_lines.append(next_text)
                break

        if not gvhd_found_para:
            # Không tìm thấy dòng GVHD trên trang bìa → chỉ cảnh báo nhẹ
            pass
        elif not gvhd_lines:
            self.result.issues.append(Issue(
                "Trang bìa — GVHD", "WARNING",
                "Tìm thấy ô 'NGƯỜI HƯỚNG DẪN' nhưng chưa điền tên giảng viên.",
                "Trang bìa",
                "Điền học vị và họ tên giảng viên hướng dẫn vào trang bìa."
            ))
        else:
            if advisor_list:
                for gvhd_raw in gvhd_lines:
                    # Bỏ qua placeholder "TS/ThS. ...", "(nếu có)", ...
                    if re.search(r'\.\.\.|nếu\s+có|ts/ths|placeholder', gvhd_raw, re.IGNORECASE):
                        continue
                    in_list = _advisor_in_list(gvhd_raw, advisor_list)
                    if not in_list:
                        short_gvhd = gvhd_raw[:60]
                        self.result.issues.append(Issue(
                            "Trang bìa — GVHD", "WARNING",
                            f"Tên GVHD không tìm thấy trong danh sách: '{short_gvhd}'",
                            "Trang bìa",
                            "Kiểm tra lại tên và học vị GVHD. Nếu đúng, hãy thêm vào config_kltn.json."
                        ))
                    else:
                        self.result.issues.append(Issue(
                            "Trang bìa — GVHD", "INFO",
                            f"GVHD xác nhận: '{gvhd_raw[:60]}'",
                            "Trang bìa", ""
                        ))
            else:
                # Chưa cấu hình danh sách → chỉ thông báo tên tìm được
                names_str = " | ".join(gvhd_lines[:3])
                self.result.issues.append(Issue(
                    "Trang bìa — GVHD", "INFO",
                    f"GVHD phát hiện: '{names_str}'. Chưa có danh sách để đối chiếu.",
                    "Trang bìa",
                    "Thêm danh sách GVHD vào file config_kltn.json để kiểm tra tính chính xác."
                ))

    # ── 1. Cài đặt trang ─────────────────────────────────────────
    def _check_page_setup(self):
        for si, sect in enumerate(self.doc.sections):
            loc = f"Section {si+1}"
            w = _emu_to_cm(sect.page_width.emu if sect.page_width else None)
            h = _emu_to_cm(sect.page_height.emu if sect.page_height else None)
            lm = _emu_to_cm(sect.left_margin.emu if sect.left_margin else None)
            rm = _emu_to_cm(sect.right_margin.emu if sect.right_margin else None)
            tm = _emu_to_cm(sect.top_margin.emu if sect.top_margin else None)
            bm = _emu_to_cm(sect.bottom_margin.emu if sect.bottom_margin else None)

            tol = Std.TOLERANCE_CM

            # Khổ giấy
            if w is not None and abs(w - Std.PAPER_W_CM) > tol:
                self.result.issues.append(Issue(
                    "Khổ giấy", "ERROR",
                    f"Chiều rộng trang {w:.2f} cm ≠ chuẩn {Std.PAPER_W_CM} cm",
                    loc, "Chỉnh khổ giấy về A4 (21 x 29.7 cm) trong Page Layout → Size."
                ))
            if h is not None and abs(h - Std.PAPER_H_CM) > tol:
                self.result.issues.append(Issue(
                    "Khổ giấy", "ERROR",
                    f"Chiều cao trang {h:.2f} cm ≠ chuẩn {Std.PAPER_H_CM} cm",
                    loc, "Chỉnh khổ giấy về A4."
                ))

            # Lề
            margin_checks = [
                (lm, Std.MARGIN_L_CM, "Lề trái"),
                (rm, Std.MARGIN_R_CM, "Lề phải"),
                (tm, Std.MARGIN_T_CM, "Lề trên"),
                (bm, Std.MARGIN_B_CM, "Lề dưới"),
            ]
            for val, expected, name in margin_checks:
                if val is not None and abs(val - expected) > tol:
                    self.result.issues.append(Issue(
                        "Lề giấy", "ERROR",
                        f"{name}: {val:.2f} cm ≠ chuẩn {expected:.2f} cm",
                        loc,
                        f"Đặt {name} = {expected} cm trong Page Layout → Margins."
                    ))

    # ── 2. Font chữ và style ─────────────────────────────────────
    def _check_font_and_styles(self):
        """Kiểm tra font Times New Roman và cỡ chữ theo từng đoạn."""
        style_errors = {}   # {style_name: count}
        font_errors  = {}
        size_errors  = {}

        SKIP_STYLES = {"Normal Table", "Default Paragraph Font", "Table Grid"}

        CONTENT_STYLES = {
            "Body LA", "Content", "Normal", "Nội dung", "Body Text",
        }
        HEADING_STYLES = {"Heading 1", "Heading 2", "Heading 3", "Heading 4"}

        for pi, para in enumerate(self._paras):
            text = para.text.strip()
            if not text or len(text) < 3:
                continue

            sname = para.style.name
            if sname in SKIP_STYLES:
                continue

            for run in para.runs:
                if not run.text.strip():
                    continue

                # Kiểm tra font
                fn = run.font.name
                if fn is None:
                    # Kế thừa từ style
                    fn = para.style.font.name
                if fn and fn not in (Std.FONT_MAIN, None):
                    key = f"{sname}:{fn}"
                    font_errors[key] = font_errors.get(key, 0) + 1

                # Kiểm tra cỡ chữ
                sz = run.font.size
                if sz is None:
                    sz = para.style.font.size
                if sz is not None:
                    sz_pt = sz / 12700
                    if sname in HEADING_STYLES:
                        pass  # kiểm tra riêng ở _check_headings
                    elif sname in CONTENT_STYLES and sz_pt < 9.5:
                        key = f"{sname}:{sz_pt:.0f}pt"
                        size_errors[key] = size_errors.get(key, 0) + 1

        # Gom báo lỗi font
        for key, count in list(font_errors.items())[:10]:
            sname, fn = key.split(':', 1)
            self.result.issues.append(Issue(
                "Font chữ", "ERROR",
                f"Font '{fn}' (không phải Times New Roman) — xuất hiện {count} lần trong style '{sname}'",
                sname,
                f"Chọn toàn bộ văn bản (Ctrl+A) và đặt font Times New Roman."
            ))

        for key, count in list(size_errors.items())[:5]:
            sname, sz = key.split(':', 1)
            self.result.issues.append(Issue(
                "Cỡ chữ", "WARNING",
                f"Cỡ chữ {sz} quá nhỏ (< 10pt) — style '{sname}' — {count} lần",
                sname,
                "Cỡ chữ nội dung tối thiểu 10pt, khuyến nghị 13pt."
            ))

    # ── 3. Cấu trúc tài liệu ─────────────────────────────────────
    def _check_structure(self):
        """Kiểm tra sự tồn tại của các phần bắt buộc."""
        all_text_lower = '\n'.join(p.text.lower() for p in self._paras)

        for section in Std.REQUIRED_SECTIONS:
            if section not in all_text_lower:
                # Map thân thiện
                display = {
                    "lời cam đoan":          "Lời cam đoan",
                    "mục lục":               "Mục lục",
                    "danh mục":              "Danh mục hình ảnh / bảng biểu",
                    "tài liệu tham khảo":    "Tài liệu tham khảo",
                }.get(section, section.title())
                self.result.issues.append(Issue(
                    "Cấu trúc", "ERROR",
                    f"Thiếu phần bắt buộc: '{display}'",
                    "",
                    f"Thêm phần '{display}' vào tài liệu theo đúng cấu trúc hướng dẫn."
                ))

    # ── 4. Heading (tên chương / tiểu mục) ───────────────────────
    def _check_headings(self):
        h1_list = []
        h2_list = []
        h3_list = []
        h4_list = []

        HEADING_MAX_SIZE = {
            "Heading 1": Std.H1_SIZE_PT,
            "Heading 2": Std.H2_SIZE_PT,
            "Heading 3": Std.H3_SIZE_PT,
            "Heading 4": Std.H4_SIZE_PT,
        }

        for pi, para in enumerate(self._paras):
            sname = para.style.name
            text = para.text.strip()
            if not text:
                continue

            if sname == "Heading 1":
                h1_list.append((pi, text))
                self._check_heading_format(para, 1)
            elif sname == "Heading 2":
                h2_list.append((pi, text))
                self._check_heading_format(para, 2)
            elif sname == "Heading 3":
                h3_list.append((pi, text))
                self._check_heading_format(para, 3)
            elif sname == "Heading 4":
                h4_list.append((pi, text))
                self._check_heading_format(para, 4)

        # Phải có ít nhất 1 Heading 1 (chương)
        if not h1_list:
            self.result.issues.append(Issue(
                "Heading", "ERROR",
                "Không có đề mục Chương (Heading 1) nào trong tài liệu.",
                "",
                "Áp dụng Style 'Heading 1' cho tên các chương."
            ))
        else:
            # Kiểm tra số lượng chương hợp lý (3–7 chương)
            if len(h1_list) < 2:
                self.result.issues.append(Issue(
                    "Cấu trúc", "WARNING",
                    f"Chỉ có {len(h1_list)} Heading 1 (Chương). KLTN thường có ít nhất 3 chương.",
                    "",
                    "Kiểm tra lại cấu trúc nội dung."
                ))

        # Kiểm tra nội dung chương (từ khoá bắt buộc)
        h1_texts = [t.lower() for _, t in h1_list]
        found_keys = []
        for kw in Std.REQUIRED_HEADING_KEYWORDS:
            for ht in h1_texts:
                if kw in ht:
                    found_keys.append(kw)
                    break

        # Không enforce cứng — chỉ warning nếu thiếu hẳn cả 3
        if len(found_keys) == 0 and len(h1_list) >= 2:
            self.result.issues.append(Issue(
                "Nội dung", "WARNING",
                "Không tìm thấy các chương điển hình (Tổng quan / Thực trạng / Giải pháp). "
                "Hãy kiểm tra tên chương có phù hợp đề cương không.",
                "",
                "Đặt tên chương rõ ràng theo đề cương của Bộ môn."
            ))

    def _check_heading_format(self, para, level: int):
        text = para.text.strip()
        loc = f"Heading {level}: '{text[:50]}'"

        # Lấy font size từ runs
        sizes = []
        bolds = []
        italics = []
        fonts = []

        for run in para.runs:
            if not run.text.strip():
                continue
            sz = run.font.size
            if sz is None:
                sz = para.style.font.size
            if sz:
                sizes.append(sz / 12700)
            bold_val = run.font.bold
            if bold_val is None:
                bold_val = para.style.font.bold
            bolds.append(bool(bold_val))
            italic_val = run.font.italic
            if italic_val is None:
                italic_val = para.style.font.italic
            italics.append(bool(italic_val))

            fn = run.font.name or para.style.font.name
            if fn:
                fonts.append(fn)

        expected_size = {1: 14, 2: 13, 3: 13, 4: 13}[level]
        if sizes:
            avg_sz = sum(sizes) / len(sizes)
            if abs(avg_sz - expected_size) > 1.0:
                self.result.issues.append(Issue(
                    "Cỡ chữ Heading", "WARNING",
                    f"Cỡ chữ {avg_sz:.0f}pt ≠ chuẩn {expected_size}pt",
                    loc,
                    f"Đặt cỡ chữ {expected_size}pt cho Heading {level}."
                ))

        # Kiểm tra Spacing và Alignment
        pf = para.paragraph_format
        sb_emu = pf.space_before if pf.space_before is not None else para.style.paragraph_format.space_before
        sa_emu = pf.space_after if pf.space_after is not None else para.style.paragraph_format.space_after
        sb_pt = _pt_from_emu(sb_emu) or 0
        sa_pt = _pt_from_emu(sa_emu) or 0
        
        ls_rule = pf.line_spacing_rule if pf.line_spacing_rule is not None else para.style.paragraph_format.line_spacing_rule
        align = pf.alignment if pf.alignment is not None else para.style.paragraph_format.alignment
        indent = pf.first_line_indent if pf.first_line_indent is not None else para.style.paragraph_format.first_line_indent
        
        # Tiêu chuẩn Spacing
        if level == 1:
            exp_sb, exp_sa = 24, 24
        else:
            exp_sb, exp_sa = 6, 12
        
        if abs(sb_pt - exp_sb) > 1.0 or abs(sa_pt - exp_sa) > 1.0:
            self.result.issues.append(Issue(
                "Heading Spacing", "WARNING",
                f"Khoảng cách đoạn chưa chuẩn: Before {sb_pt:.0f}pt (chuẩn {exp_sb}pt), After {sa_pt:.0f}pt (chuẩn {exp_sa}pt).",
                loc,
                f"Sửa Paragraph Spacing cho Heading {level} thành Before {exp_sb}pt, After {exp_sa}pt."
            ))
            
        # Line spacing = single
        if ls_rule is not None and ls_rule.value != 0: # 0 = WD_LINE_SPACING.SINGLE
            self.result.issues.append(Issue(
                "Heading Spacing", "WARNING",
                f"Giãn dòng của Heading {level} không phải Single.",
                loc,
                f"Sửa Line spacing thành Single."
            ))
            
        # Alignment & Indent
        if level == 1:
            # Heading 1: căn giữa (CENTER) HOẶC trái (LEFT) đều OK
            # Không thụt đầu dòng
            if align is not None and align not in (
                WD_ALIGN_PARAGRAPH.CENTER,
                WD_ALIGN_PARAGRAPH.LEFT,
            ):
                self.result.issues.append(Issue(
                    "Heading Alignment", "WARNING",
                    f"Canh lề Heading 1 phải là Giữa (Center) hoặc Trái (Left), hiện đang khác.",
                    loc, "Đặt canh lề Center hoặc Left cho tên chương (Heading 1)."
                ))
            if indent and indent > 0:
                self.result.issues.append(Issue(
                    "Heading Indent", "WARNING",
                    "Heading 1 bị thụt đầu dòng — tên chương không được thụt lề.",
                    loc, "Đặt First line indent = 0 cho Heading 1."
                ))
        else:
            # Heading 2/3/4: canh trái
            if align is not None and align != WD_ALIGN_PARAGRAPH.LEFT:
                self.result.issues.append(Issue(
                    "Heading Alignment", "WARNING",
                    f"Canh lề Heading {level} không phải trái (Left).",
                    loc, "Canh lề trái (Left) cho tên tiểu mục."
                ))
            if indent and indent > 0:
                self.result.issues.append(Issue(
                    "Heading Indent", "WARNING",
                    f"Heading {level} bị thụt đầu dòng.",
                    loc, "Không thụt đầu dòng (First line indent = 0) cho tên tiểu mục."
                ))

        # Heading 1: phải in đậm, IN HOA
        if level == 1:
            if bolds and not all(bolds):
                self.result.issues.append(Issue(
                    "Heading", "WARNING",
                    "Tên chương (Heading 1) phải in đậm (Bold).",
                    loc, "Bôi đen tên chương và bấm Ctrl+B."
                ))
            if text and not text == text.upper():
                alpha_chars = [c for c in text if c.isalpha()]
                if alpha_chars and not all(c.isupper() for c in alpha_chars):
                    self.result.issues.append(Issue(
                        "Heading", "WARNING",
                        f"Tên chương (Heading 1) không in hoa.",
                        loc, "Heading 1 phải viết IN HOA toàn bộ (Vd: CHƯƠNG 1)."
                    ))

        # Heading 2: phải in đậm
        elif level == 2:
            if bolds and not all(bolds):
                self.result.issues.append(Issue(
                    "Heading", "WARNING",
                    "Tiểu mục 1 (Heading 2) phải in đậm.",
                    loc, "Bôi đen và Ctrl+B."
                ))

        # Heading 3: in đậm + nghiêng
        elif level == 3:
            if bolds and not all(bolds):
                self.result.issues.append(Issue(
                    "Heading", "WARNING",
                    "Tiểu mục 2 (Heading 3) phải in đậm.",
                    loc, "Bôi đen và bấm Ctrl+B."
                ))
            if italics and not all(italics):
                self.result.issues.append(Issue(
                    "Heading", "WARNING",
                    "Tiểu mục 2 (Heading 3) phải in nghiêng.",
                    loc, "Bôi đen và bấm Ctrl+I."
                ))

        # Heading 4: in nghiêng
        elif level == 4:
            if italics and not all(italics):
                self.result.issues.append(Issue(
                    "Heading", "WARNING",
                    "Tiểu mục 3 (Heading 4) phải in nghiêng.",
                    loc, "Bôi đen và bấm Ctrl+I."
                ))

    # ── 5. Nội dung đoạn văn ─────────────────────────────────────
    def _check_body_text(self):
        """Kiểm tra style 'Body LA', 'Content', 'Normal' ở phần nội dung."""
        BODY_STYLES = {"Body LA", "Content", "Normal", "Body Text"}

        body_paras = [p for p in self._paras
                      if p.style.name in BODY_STYLES and len(p.text.strip()) > 20]

        if not body_paras:
            self.result.issues.append(Issue(
                "Nội dung", "WARNING",
                "Không tìm thấy đoạn văn nội dung nào với style chuẩn (Body LA / Content).",
                "",
                "Áp dụng style 'Body LA' hoặc 'Content' cho toàn bộ phần nội dung."
            ))
            return

        # Mẫu kiểm tra: lấy 20 đoạn đại diện
        sample = body_paras[: min(30, len(body_paras))]

        wrong_spacing = 0
        wrong_align   = 0
        wrong_indent  = 0
        wrong_font    = 0

        for para in sample:
            pf = para.paragraph_format
            ls = pf.line_spacing
            ls_rule = pf.line_spacing_rule

            # Nếu para không có line_spacing, kế thừa từ style
            if ls is None:
                ls = para.style.paragraph_format.line_spacing
                ls_rule = para.style.paragraph_format.line_spacing_rule

            # Tính Spacing before/after
            sb_emu = pf.space_before if pf.space_before is not None else para.style.paragraph_format.space_before
            sa_emu = pf.space_after if pf.space_after is not None else para.style.paragraph_format.space_after
            sb_pt = _pt_from_emu(sb_emu) or 0
            sa_pt = _pt_from_emu(sa_emu) or 0
            
            # Check spacing Before 10pt, After 0pt, Line spacing 1.5
            is_15 = False
            if ls is not None:
                if ls_rule and ls_rule.value == 1:   # ONE_POINT_FIVE
                    is_15 = True
                elif ls_rule is None or ls_rule.value == 5:  # MULTIPLE
                    if isinstance(ls, float) and abs(ls - 1.5) < 0.1:
                        is_15 = True
                    elif isinstance(ls, int) and abs(ls / 12700 / 12 - 1.5) < 0.1:
                        is_15 = True
            
            if not is_15 or abs(sb_pt - 10) > 1.0 or abs(sa_pt - 0) > 1.0:
                wrong_spacing += 1

            # Justify (canh đều 2 bên) = 3
            al = para.alignment
            if al is None:
                al = para.style.paragraph_format.alignment
            if al is not None and al != WD_ALIGN_PARAGRAPH.JUSTIFY:
                wrong_align += 1
                
            # Không thụt đầu dòng (indent = 0)
            ind = pf.first_line_indent if pf.first_line_indent is not None else para.style.paragraph_format.first_line_indent
            if ind and ind > 0:
                wrong_indent += 1
                
            # Font TNR, 13pt
            sz = None
            fn = None
            if para.runs:
                # Find first run with explicit font
                for r in para.runs:
                    if r.text.strip():
                        sz = _pt_from_emu(r.font.size)
                        fn = r.font.name
                        if sz or fn:
                            break
            if sz is None:
                sz = _pt_from_emu(para.style.font.size)
            if fn is None:
                fn = para.style.font.name
            
            if (sz and abs(sz - 13.0) > 0.5) or (fn and fn != "Times New Roman"):
                wrong_font += 1

        if wrong_spacing > len(sample) * 0.4:
            self.result.issues.append(Issue(
                "Nội dung - Spacing", "ERROR",
                f"{wrong_spacing}/{len(sample)} đoạn mẫu có khoảng cách / giãn dòng sai (Chuẩn: Before 10, After 0, Line spacing 1.5).",
                "Body text",
                "Đặt thông số khoảng cách cho dòng nội dung: Before 10pt, After 0pt, 1.5 lines."
            ))

        if wrong_align > len(sample) * 0.4:
            self.result.issues.append(Issue(
                "Nội dung - Alignment", "ERROR",
                f"{wrong_align}/{len(sample)} đoạn mẫu không canh chỉ đều hai bên (Justify).",
                "Body text",
                "Cần chọn toàn bộ nội dung (Ctrl+A) và chọn canh đều hai bên (Ctrl+J)."
            ))
            
        if wrong_indent > len(sample) * 0.4:
            self.result.issues.append(Issue(
                "Nội dung - Thụt dòng", "WARNING",
                f"{wrong_indent}/{len(sample)} đoạn bị thụt lề đầu dòng (First line indent > 0).",
                "Body text",
                "Đoạn văn KLTN tiêu chuẩn không được thụt đầu dòng."
            ))
            
        if wrong_font > len(sample) * 0.4:
            self.result.issues.append(Issue(
                "Nội dung - Font", "WARNING",
                f"{wrong_font}/{len(sample)} đoạn chưa đúng font Times New Roman hoặc chưa là cỡ 13pt.",
                "Body text",
                "Đổi định dạng toàn bộ nội dung thành font Times New Roman, cỡ 13pt."
            ))

    # ── 6. Caption (chú thích hình / bảng) ───────────────────────
    def _check_captions(self):
        """Kiểm tra caption tồn tại và đúng định dạng."""
        caption_paras = [p for p in self._paras if p.style.name == "Caption"]

        if not caption_paras:
            # Thử tìm bằng text pattern
            fig_cap = [p for p in self._paras
                       if re.search(r'^(Hình|Bảng|Figure|Table)\s+\d', p.text.strip())]
            if not fig_cap:
                self.result.issues.append(Issue(
                    "Chú thích", "WARNING",
                    "Không tìm thấy caption (Hình x.y / Bảng x.y) nào trong tài liệu.",
                    "",
                    "Thêm chú thích cho tất cả hình và bảng biểu bằng style 'Caption'."
                ))
                return

        # Kiểm tra định dạng "Hình x.y" hay "Bảng x.y"
        # Chấp nhận cả placeholder 'Hình .:' (file mẫu có thể dùng kiểu SEQ field)
        wrong_fmt = []
        for p in caption_paras:
            t = p.text.strip()
            # OK nếu bắt đầu bằng Hình/Bảng/Figure/Table + (số hoặc dấu chấm)
            if not re.match(
                r'^(Hình|Bảng|Figure|Table|Hình vẽ)[\s\xa0]+[\d.:]',
                t, re.IGNORECASE
            ):
                wrong_fmt.append(t[:60])

        if len(wrong_fmt) > max(3, len(caption_paras) * 0.5):
            self.result.issues.append(Issue(
                "Chú thích", "WARNING",
                f"{len(wrong_fmt)}/{len(caption_paras)} caption không đúng định dạng 'Hình x.y' / 'Bảng x.y'.",
                "Caption",
                "Đặt caption theo định dạng: 'Hình 2.1 Tiêu đề hình' phía dưới hình; "
                "'Bảng 2.1 Tiêu đề bảng' phía trên bảng. Dùng Insert → Caption để tự động đánh số."
            ))

        # Kiểm tra font size, align, spacing
        non_italic = 0
        wrong_align_captions = 0
        wrong_spacing_captions = 0
        
        sample_caps = caption_paras[: min(20, len(caption_paras))]
        for p in sample_caps:
            # Kiểm tra Italic, Size
            is_italic = p.style.font.italic
            sz = _pt_from_emu(p.style.font.size)
            fn = p.style.font.name
            
            for run in p.runs:
                if run.text.strip():
                    if run.font.italic is not None:
                        is_italic = run.font.italic
                    if run.font.size is not None:
                        sz = _pt_from_emu(run.font.size)
                    if run.font.name is not None:
                        fn = run.font.name
                        
            if not is_italic or (sz and abs(sz - 12.0) > 0.5):
                non_italic += 1
                
            # Alignment: CENTER = 1
            al = p.alignment if p.alignment is not None else p.style.paragraph_format.alignment
            if al is not None and al != WD_ALIGN_PARAGRAPH.CENTER:
                wrong_align_captions += 1
                
            # Spacing Before = 6pt, After = 6pt, Line spacing = Single
            pf = p.paragraph_format
            sb_emu = pf.space_before if pf.space_before is not None else p.style.paragraph_format.space_before
            sa_emu = pf.space_after if pf.space_after is not None else p.style.paragraph_format.space_after
            sb_pt = _pt_from_emu(sb_emu) or 0
            sa_pt = _pt_from_emu(sa_emu) or 0
            ls_rule = pf.line_spacing_rule if pf.line_spacing_rule is not None else p.style.paragraph_format.line_spacing_rule
            
            if abs(sb_pt - 6.0) > 1.0 or abs(sa_pt - 6.0) > 1.0 or (ls_rule is not None and ls_rule.value != 0):
                wrong_spacing_captions += 1

        if non_italic > max(2, len(sample_caps) * 0.3):
            self.result.issues.append(Issue(
                "Caption - Font", "WARNING",
                f"{non_italic}/{len(sample_caps)} caption bị thiết lập sai font (chuẩn: Times New Roman, 12pt, in nghiêng).",
                "Caption",
                "Đổi toàn bộ định dạng caption: 12pt, in nghiêng, Times New Roman."
            ))
            
        if wrong_align_captions > max(2, len(sample_caps) * 0.3):
            self.result.issues.append(Issue(
                "Caption - Canh lề", "WARNING",
                f"{wrong_align_captions}/{len(sample_caps)} caption không được canh lề chính giữa (Center).",
                "Caption",
                "Toàn bộ caption Hình/Bảng biểu phải được canh giữa."
            ))
            
        if wrong_spacing_captions > max(2, len(sample_caps) * 0.3):
            self.result.issues.append(Issue(
                "Caption - Spacing", "WARNING",
                f"{wrong_spacing_captions}/{len(sample_caps)} caption khoảng cách không đúng (Chuẩn: Before 6pt, After 6pt, Single).",
                "Caption",
                "Sửa Line/Paragraph spacing của Caption thành Before 6pt, After 6pt, Single."
            ))

    # ── 7. Tài liệu tham khảo ─────────────────────────────────────
    def _check_references(self):
        """Kiểm tra sự tồn tại và định dạng cơ bản của TLTK.
        Chấp nhận cả chuẩn IEEE ([1]) và APA (Tác giả, năm).
        """
        full_text = '\n'.join(p.text for p in self._paras)

        ref_idx = None
        # Tìm ngược từ dưới lên để bắt đúng Tiêu đề (không bị nhầm với Mục lục)
        for i in range(len(self._paras)-1, -1, -1):
            text = self._paras[i].text.strip().lower()
            if 'tài liệu tham khảo' in text and len(text) < 40:
                ref_idx = i
                break

        if ref_idx is None:
            self.result.issues.append(Issue(
                "Tài liệu tham khảo", "ERROR",
                "Không tìm thấy phần 'TÀI LIỆU THAM KHẢO'.",
                "",
                "Thêm phần Tài liệu tham khảo ở cuối tài liệu."
            ))
            return

        # Đếm số tài liệu tham khảo
        ref_count = 0
        for p in self._paras[ref_idx + 1:]:
            t = p.text.strip()
            if not t: continue
            
            # Gặp phụ lục => dừng
            if 'phụ lục' in t.lower() and len(t) < 40:
                break
                
            ref_count += 1

        if ref_count < 3:
            self.result.issues.append(Issue(
                "Tài liệu tham khảo", "WARNING",
                f"Số lượng tài liệu tham khảo rất ít: {ref_count} tài liệu.",
                "Tài liệu tham khảo",
                "KLTN thường cần ít nhất 5–10 tài liệu tham khảo."
            ))
        else:
            self.result.issues.append(Issue(
                "Tài liệu tham khảo", "INFO",
                f"Tìm thấy khoảng {ref_count} tài liệu tham khảo.",
                "Tài liệu tham khảo", ""
            ))

        # ── Phát hiện chuẩn trích dẫn (chỉ tìm trong phần nội dung) ──
        content_text = '\n'.join(p.text for p in self._paras[:ref_idx])

        # IEEE: [1], [2], [1,2], [1-3], [1, 2, 3]
        ieee_count = len(re.findall(r'\[\d+(?:[,\-\u2013]\s*\d+)*\]', content_text))

        # APA in-text — hỗ trợ 2 dạng:
        #  1. Parenthetical: (Nguyen, 2023), (Smith & Jones, 2019), (Trần et al., 2023)
        #                    (World Bank, 2023), (Bộ Tài chính, 2022)
        #  2. Narrative:     Nguyen (2023), Smith et al. (2022), Trần Văn A (2021)
        apa_parenthetical = len(re.findall(
            r'\([A-Z\u00C0-\u024F\u1EA0-\u1EF9][^\(\)]{2,60},\s*(?:19|20)\d{2}[a-z]?\)',
            content_text
        ))
        apa_narrative = len(re.findall(
            r'(?<!\[)[A-Z\u00C0-\u024F\u1EA0-\u1EF9]'   # chữ cái đầu viết hoa
            r'[\wÀ-ỹ\s\.\&]{3,50}'                       # tên tác giả
            r'\s+\((?:19|20)\d{2}[a-z]?\)',               # (năm)
            content_text
        ))
        apa_count = apa_parenthetical + apa_narrative

        has_citation = ieee_count > 0 or apa_count > 0

        if not has_citation:
            self.result.issues.append(Issue(
                "Trích dẫn", "WARNING",
                "Không tìm thấy trích dẫn trong nội dung.\n"
                "  · Chuẩn IEEE: [1], [2,3], [1-4]\n"
                "  · Chuẩn APA : (Tác giả, năm) ví dụ (Nguyen, 2023), (Smith et al., 2022)",
                "",
                "Thêm trích dẫn ngay sau mỗi thông tin tham khảo. "
                "Chấp nhận cả IEEE và APA — chọn một chuẩn và dùng nhất quán."
            ))
            return

        # Xác định chuẩn chính đang dùng
        if ieee_count >= apa_count:
            primary_style, primary_count = "IEEE", ieee_count
        else:
            primary_style, primary_count = "APA", apa_count

        mix_note = ""
        minor_count = min(ieee_count, apa_count)
        if ieee_count > 0 and apa_count > 0:
            other = "APA" if primary_style == "IEEE" else "IEEE"
            mix_note = f" | cũng có {minor_count} trích dẫn kiểu {other}"

        self.result.issues.append(Issue(
            "Trích dẫn", "INFO",
            f"Chuẩn trích dẫn: {primary_style} — {primary_count} lần{mix_note}",
            "Trích dẫn", ""
        ))

        # Cảnh báo nếu dùng lẫn lộn 2 chuẩn quá nhiều (> 3 trường hợp chuẩn phụ)
        MIX_THRESHOLD = 3
        if ieee_count > 0 and apa_count > 0 and minor_count > MIX_THRESHOLD:
            self.result.issues.append(Issue(
                "Trích dẫn", "WARNING",
                f"Lẫn lộn 2 chuẩn: IEEE ({ieee_count} lần) và APA ({apa_count} lần).",
                "Trích dẫn",
                "Chọn MỘT chuẩn duy nhất và áp dụng nhất quán toàn tài liệu. "
                "Kỹ thuật → IEEE  |  Kinh tế / Quản lý → APA."
            ))

    # ── 8. Chữ viết tắt và trích dẫn trực tiếp ────────────────────
    def _check_abbreviations_and_quotes(self):
        """Kiểm tra tuân thủ dùng viết tắt và cách trích dẫn nguyên văn."""
        full_text = '\n'.join(p.text for p in self._paras)
        
        # 1. Từ viết tắt
        # Tìm danh mục từ viết tắt ở phần đầu (khoảng 4000 ký tự đầu tiên hoặc 80 đoạn)
        head_text = '\n'.join(p.text for p in self._paras[:100]).lower()
        has_abbrev_list = bool(re.search(r'danh\s+mục\s+từ\s+viết\s+tắt|danh\s+mục\s+chữ\s+viết\s+tắt|bảng\s+viết\s+tắt|bảng\s+từ\s+viết\s+tắt|ký\s+hiệu\s+viết\s+tắt', head_text))
        
        # Chỉ check chữ cái thuần (tránh các từ tiếng Việt có dấu ALL CAPS)
        abbrevs = set(re.findall(r'\b[A-Z]{2,14}\b', full_text))
        ignore_words = {'CHƯỜNG', 'BẢNG', 'HÌNH', 'LỜI', 'CAM', 'ĐOAN', 'MỤC', 'LỤC', 'DANH', 'TÀI', 'LIỆU', 'THAM', 'KHẢO',
          'PHẦN', 'KẾT', 'LUẬN', 'TỔNG', 'QUAN', 'THỰC', 'TRẠNG', 'GIẢI', 'PHÁP', 'KÍNH', 'CHÚC', 'XIN', 'NGƯỜI', 'HƯỚNG', 'DẪN', 'SINH', 'VIÊN', 'THỰC', 'HIỆN', 'TRƯỜNG', 'ĐẠI', 'HỌC', 'THỦY', 'LỢI', 'BỘ', 'GIÁO', 'DỤC', 'ĐÀO', 'TẠO', 'NÔNG', 'NGHIỆP', 'MÔI', 'TRƯỜNG', 'PHỤ', 'CÁC', 'VÀ', 'SỰ', 'CHO', 'KHÔNG', 'LÀ', 'CỦA', 'ĐẠT'}
        abbrevs = {a for a in abbrevs if a not in ignore_words and sum(c.isalpha() for c in a) == len(a)}
        
        if len(abbrevs) > 10 and not has_abbrev_list:
            abbrev_sample = ', '.join(list(abbrevs)[:5])
            self.result.issues.append(Issue(
                "Viết tắt", "WARNING",
                f"Sử dụng nhiều từ viết tắt ({len(abbrevs)} từ, VD: {abbrev_sample}) nhưng không có Danh mục ký hiệu/viết tắt.",
                "Phần mở đầu",
                "Nếu có quá nhiều chữ viết tắt thì phải có Bảng danh mục các từ viết tắt (xếp theo thứ tự A, B, C) ở phần đầu ĐATN/KLTN."
            ))
            
        # 2. Quotes (Trích dẫn trực tiếp)
        # Bắt buộc nếu có "ngoặc kép nguyên văn" => phải có dẫn nguồn & ko đc quá dài.
        BODY_STYLES = {"Body LA", "Content", "Normal", "Body Text"}
        content_paras = [p for p in self._paras if p.style.name in BODY_STYLES and len(p.text.strip()) > 10]
        
        # Giới hạn cảnh báo cho ngoặc kép (tránh spam nều tài liệu trích quá nhiều)
        warned_quotes = 0
        for p in content_paras:
            if warned_quotes >= 5: break
            
            text = p.text
            quotes = re.findall(r'["“«]([^"”»]{15,})["”»]', text) # chỉ quan tâm quote dài hơn 15 ký tự
            for q in quotes:
                words_count = len(q.split())
                # Dấu hiệu cite trong cung paragraph (APA hoăc IEEE)
                has_cite = bool(re.search(r'\[\d+(?:[,\-\u2013]\s*\d+)*\]|\([A-Z\u00C0-\u1EF9].+?,\s*(?:19|20)\d{2}[a-z]?\)', text))
                
                # Check nếu quote chứa nhiều hơn 2 câu
                sentences = re.split(r'[.!?]+', q)
                real_sentences = [s for s in sentences if len(s.strip()) > 5]
                
                if words_count > 60 or len(real_sentences) > 2:
                    self.result.issues.append(Issue(
                        "Trích dẫn", "WARNING",
                        f"Ghi nguyên văn quá dài ({len(real_sentences)} câu, {words_count} từ): '\"{q[:40]}...\"'",
                        f"Đoạn: '{text[:20]}...'",
                        "Chỉ được ghi nguyên văn một hoặc hai câu. Ghi nguyên văn đoạn dài là đạo văn, hãy viết lại bằng ngôn từ của mình (Paragraphing)."
                    ))
                    warned_quotes += 1
                elif not has_cite:
                    self.result.issues.append(Issue(
                        "Trích dẫn", "WARNING",
                        f"Trích dẫn nguyên văn nhưng không ghi nguồn: '\"{q[:40]}...\"'",
                        f"Đoạn: '{text[:20]}...'",
                        "Khi ghi lại nguyên văn phải đặt trong ngoặc kép VÀ bắt buộc ghi nguồn trích dẫn ngay sau đó (IEEE / APA)."
                    ))
                    warned_quotes += 1

    # ── 9. Phát hiện nghi vấn copy từ AI / lỗi copy-paste ────────
    def _check_ai_copy_anomalies(self):
        """Kiểm tra các dấu hiệu copy paste từ AI (Markdown, nhiều cụm từ bold bất thường, text in hoa)."""
        BODY_STYLES = {"Body LA", "Content", "Normal", "Body Text"}
        
        markdown_bold_count = 0
        all_caps_count = 0
        ai_pattern_bold_count = 0
        ai_pronoun_count = 0
        ai_mid_bold_paragraphs = 0
        
        passed_chapter_1 = False
        
        for p in self._paras:
            text = p.text.strip()
            if not text:
                continue
                
            text_lower = text.lower()
            
            # Xác định vị trí bắt đầu phần nội dung chính (Chương 1)
            if not passed_chapter_1 and p.style.name == "Heading 1" and "chương 1" in text_lower:
                passed_chapter_1 = True
                
            if p.style.name not in BODY_STYLES or len(text) <= 10:
                continue
            
            # 1. Phát hiện dấu ** trong đoạn văn
            if "**" in text:
                markdown_bold_count += 1
                
            # 2. Phát hiện chữ in hoa nguyên khối dài (trên 3 từ, trên 15 kí tự) ko phải format chuẩn
            # (Tránh nhầm với chữ chương, bảng...)
            caps = re.findall(r'\b[A-Z\u00C0-\u1EF9\s]{15,}\b', text)
            for c in caps:
                trim_c = c.strip()
                if len(trim_c.split()) >= 3 and sum(ch.isalpha() for ch in trim_c) > 10:
                    # Ignore common titles that might just be badly formatted
                    if any(w in trim_c for w in ["CHƯƠNG", "PHẦN", "BẢNG", "HÌNH", "LỜI"]):
                        continue
                    all_caps_count += 1
                    
            # 3. Phát hiện bold đầu câu kết thúc bằng dấu hai chấm (VD: **Tính năng:** Nội dung)
            is_bold_start = False
            if p.runs and p.runs[0].font.bold:
                bold_text = p.runs[0].text.strip()
                if 2 < len(bold_text) < 40 and bold_text.endswith(':'):
                    ai_pattern_bold_count += 1
                    is_bold_start = True
                    
            # 4. Phát hiện đại từ nhân xưng bạn/tôi (dấu hiệu AI) từ chương 1 trở đi
            if passed_chapter_1:
                if re.search(r'\b(bạn|tôi)\b', text_lower):
                    ai_pronoun_count += 1

            # 5. Phát hiện nhiều cụm từ bôi đậm nằm giữa đoạn văn (nhấn mạnh AI)
            mid_bold_count = 0
            for i, r in enumerate(p.runs):
                if is_bold_start and i == 0:
                    continue  # Bỏ qua cụm in đậm đầu câu ở rule 3
                # Chỉ bắt in đậm thực sự trong run dài hơn 3 ký tự
                if r.font.bold and len(r.text.strip()) > 3:
                    mid_bold_count += 1
            if mid_bold_count >= 3:
                ai_mid_bold_paragraphs += 1
                
        if markdown_bold_count >= 2:
            self.result.issues.append(Issue(
                "Nghi vấn Copy AI", "ERROR",
                f"Phát hiện {markdown_bold_count} đoạn nội dung chứa dấu '**' (đặc trưng markdown của ChatGPT/Claude).",
                "Nội dung (Body)",
                "Xong bài nhớ xóa các dấu '**' dư thừa; chỉnh sửa lại định dạng thủ công đàng hoàng nhé."
            ))
            
        if all_caps_count >= 3:
            self.result.issues.append(Issue(
                "Nghi vấn Copy AI / Format", "WARNING",
                f"Phát hiện {all_caps_count} cụm từ/đoạn viết IN HOA toàn bộ nằm lạc lõng giữa các dòng nội dung.",
                "Nội dung (Body)",
                "Trình bày đoạn văn KHÔNG BÔI ĐEN VÀ VIẾT HOA ngẫu nhiên. Copy từ AI hay bị lỗi này. Cần định dạng lại."
            ))
            
        if ai_pattern_bold_count >= 4:
            self.result.issues.append(Issue(
                "Nghi vấn Copy AI", "WARNING",
                f"Có chứa {ai_pattern_bold_count} dòng nội dung bắt đầu bằng In đậm (Bold) kèm dấu hai chấm ':' (Vd: 'Lợi ích:').",
                "Nội dung (Body)",
                "Dấu hiệu văn phong rập khuôn của AI. Hãy viết lại mạch lạc hơn hoặc tạo Heading chuẩn nếu thực sự là các mục lục."
            ))

        if ai_mid_bold_paragraphs >= 2:
            self.result.issues.append(Issue(
                "Nghi vấn Copy AI", "WARNING",
                f"Phát hiện {ai_mid_bold_paragraphs} đoạn văn có bôi đậm (Bold) nhiều cụm từ rải rác ở giữa câu.",
                "Nội dung (Body)",
                "AI thường bôi đậm các từ khóa quan trọng ở giữa câu. Viết KLTN chuẩn không nên bôi đen lung tung, chỉ bôi đen Heading hoặc tên Hình/Bảng."
            ))

        if ai_pronoun_count > 0:
            self.result.issues.append(Issue(
                "Nghi vấn Copy AI", "WARNING",
                f"Phát hiện {ai_pronoun_count} đoạn văn có sử dụng đại từ nhân xưng 'bạn' hoặc 'tôi'.",
                "Nội dung (từ Chương 1 trở đi)",
                "Văn phong học thuật không dùng 'bạn' hay 'tôi'. Đây thường là dấu hiệu sử dụng AI dịch/tạo bài. Cần tránh và thay thế bằng từ ngữ khách quan hơn (ví dụ: 'người dùng', 'nhóm tác giả')."
            ))

    # ── 10. Tính điểm ─────────────────────────────────────────────
    def _compute_score(self):
        errors   = [i for i in self.result.issues if i.severity == "ERROR"]
        warnings = [i for i in self.result.issues if i.severity == "WARNING"]

        self.result.error_count = len(errors)
        self.result.warn_count  = len(warnings)

        # Tổng điểm trừ tối đa
        deduct = len(errors) * 10 + len(warnings) * 3
        self.result.score = max(0, 100 - deduct)


# ════════════════════════════════════════════════════════════════
#  QUÉT NHIỀU FILE
# ════════════════════════════════════════════════════════════════
def check_file(filepath: str) -> CheckResult:
    try:
        checker = KLTNChecker(filepath)
        return checker.check_all()
    except Exception as e:
        r = CheckResult(filepath=filepath)
        r.issues.append(Issue(
            "Hệ thống", "ERROR",
            f"Không thể đọc file: {e}",
            str(filepath), "Kiểm tra file có bị hỏng hay đang mở không."
        ))
        r.error_count = 1
        r.score = 0
        return r


def scan_directory(root_path: str) -> list:
    root = Path(root_path)
    docx_files = [f for f in root.rglob("*.docx")
                  if not f.name.startswith("~$")]
    results = []
    total = len(docx_files)
    for i, fp in enumerate(sorted(docx_files), 1):
        print(f"  [{i:03d}/{total}] Kiểm tra: {fp.name[:70]}")
        results.append(check_file(str(fp)))
    return results


# ════════════════════════════════════════════════════════════════
#  XUẤT KẾT QUẢ EXCEL
# ════════════════════════════════════════════════════════════════
def export_excel(results: list, output_path: str):
    wb = Workbook()

    thin = Side(style='thin', color='CCCCCC')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    thick_bot = Border(
        left=thin, right=thin,
        top=thin, bottom=Side(style='medium', color='888888')
    )

    def hcell(ws, r, c, val, bg='1F3864', fg='FFFFFF', sz=11, bold=True, wrap=True, center=True):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = XFont(name='Arial', size=sz, bold=bold, color=fg)
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(
            horizontal='center' if center else 'left',
            vertical='center', wrap_text=wrap
        )
        cell.border = bdr
        return cell

    def dcell(ws, r, c, val, bg='FFFFFF', fg='000000', sz=10, bold=False,
              center=False, italic=False, wrap=True):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = XFont(name='Arial', size=sz, bold=bold, color=fg, italic=italic)
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(
            horizontal='center' if center else 'left',
            vertical='center', wrap_text=wrap
        )
        cell.border = bdr
        return cell

    SEV_COLOR = {"ERROR": "C00000", "WARNING": "E2681A", "INFO": "196B24"}
    SEV_BG    = {"ERROR": "FFF0F0", "WARNING": "FFF8F0", "INFO": "F0FFF4"}

    # ─────────────────────────────────────────────────────────────
    # Sheet 1: Tóm tắt tổng hợp
    # ─────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Tổng hợp"
    ws1.sheet_properties.tabColor = "1F3864"

    # Tiêu đề
    ws1.merge_cells('A1:K1')
    c = ws1['A1']
    c.value = "KẾT QUẢ KIỂM TRA ĐỊNH DẠNG KHÓA LUẬN TỐT NGHIỆP"
    c.font = XFont(name='Arial', size=14, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='1F3864')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells('A2:K2')
    ts_cell = ws1['A2']
    ts_cell.value = (
        f"Xuất lúc: {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  "
        f"Tổng file kiểm tra: {len(results)}  ·  "
        f"File đạt (≥70đ): {sum(1 for r in results if r.score >= 70)}"
    )
    ts_cell.font = XFont(name='Arial', size=10, italic=True, color='555555')
    ts_cell.fill = PatternFill('solid', fgColor='F0F4FA')
    ts_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[2].height = 20

    headers = ['STT', 'Tên file', 'Sinh viên', 'MSSV', 'GVHD',
               'Lỗi nghiêm trọng', 'Cảnh báo', 'Điểm hệ 100', 'Điểm chữ', 'Đánh giá', 'Ghi chú nhanh']
    ws1.row_dimensions[3].height = 32
    for ci, h in enumerate(headers, 1):
        hcell(ws1, 3, ci, h, sz=10)

    SCORE_EVAL = {
        (90, 101): ("✅ Đạt tốt",    "00703A", "D9F2E6"),
        (70,  90): ("✔ Đạt",         "196B24", "EDF7F0"),
        (50,  70): ("⚠ Cần sửa",    "7B4A00", "FFF3E0"),
        ( 0,  50): ("❌ Không đạt", "C00000", "FFE8E8"),
    }

    for i, res in enumerate(results):
        r = i + 4
        ws1.row_dimensions[r].height = 22

        eval_text, fg, bg = "⚠ Cần sửa", "7B4A00", "FFF3E0"
        for (lo, hi), (txt, f, b) in SCORE_EVAL.items():
            if lo <= res.score < hi:
                eval_text, fg, bg = txt, f, b
                break

        alt = 'F7F9FC' if i % 2 == 0 else 'FFFFFF'
        errors_sum = [iss for iss in res.issues if iss.severity == 'ERROR']
        quick_note = '; '.join(iss.message[:40] for iss in errors_sum[:2])

        dcell(ws1, r, 1, i+1,                         bg=alt, center=True, bold=True)
        dcell(ws1, r, 2, Path(res.filepath).name,     bg=alt, sz=9)
        dcell(ws1, r, 3, res.student_name,             bg=alt)
        dcell(ws1, r, 4, res.student_id,               bg=alt, center=True)
        dcell(ws1, r, 5, res.advisor,                  bg=alt)
        dcell(ws1, r, 6, res.error_count,              bg='FFF0F0' if res.error_count else alt,
              fg='C00000' if res.error_count else '000000', center=True, bold=bool(res.error_count))
        dcell(ws1, r, 7, res.warn_count,               bg='FFF8F0' if res.warn_count else alt,
              fg='E07B00' if res.warn_count else '000000', center=True)
        dcell(ws1, r, 8, res.score,                    bg=bg, fg=fg, center=True, bold=True, sz=12)
        dcell(ws1, r, 9, res.letter_grade,             bg=bg, fg=fg, center=True, bold=True, sz=12)
        dcell(ws1, r, 10, eval_text,                   bg=bg, fg=fg, center=True, bold=True)
        dcell(ws1, r, 11, quick_note,                  bg=alt, sz=9, italic=True, fg='666666')

    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 38
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 22
    ws1.column_dimensions['F'].width = 8
    ws1.column_dimensions['G'].width = 8
    ws1.column_dimensions['H'].width = 12
    ws1.column_dimensions['I'].width = 10
    ws1.column_dimensions['J'].width = 14
    ws1.column_dimensions['K'].width = 45

    ws1.auto_filter.ref = 'A3:K3'
    ws1.freeze_panes   = 'A4'

    # ─────────────────────────────────────────────────────────────
    # Sheet 2: Chi tiết từng lỗi
    # ─────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Chi tiết lỗi")
    ws2.sheet_properties.tabColor = "C00000"

    ws2.merge_cells('A1:G1')

    c = ws2['A1']
    c.value = "CHI TIẾT CÁC LỖI VÀ CẢNH BÁO TỪNG FILE"
    c.font = XFont(name='Arial', size=13, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='7B1212')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 30

    hdrs2 = ['STT file', 'Tên file', 'Mức độ', 'Nhóm lỗi', 'Mô tả lỗi', 'Vị trí', 'Gợi ý sửa']
    for ci, h in enumerate(hdrs2, 1):
        hcell(ws2, 2, ci, h, sz=10, bg='922B21')
    ws2.row_dimensions[2].height = 26

    row = 3
    for fi, res in enumerate(results):
        issues_sorted = sorted(res.issues,
                               key=lambda x: {"ERROR": 0, "WARNING": 1, "INFO": 2}[x.severity])
        for iss in issues_sorted:
            ws2.row_dimensions[row].height = 36
            fg_c = SEV_COLOR.get(iss.severity, '000000')
            bg_c = SEV_BG.get(iss.severity, 'FFFFFF')
            alt  = 'F7F7F7' if fi % 2 == 0 else 'FFFFFF'

            dcell(ws2, row, 1, fi+1,                            bg=alt, center=True, bold=True)
            dcell(ws2, row, 2, Path(res.filepath).name,         bg=alt, sz=9)
            dcell(ws2, row, 3, iss.severity,                    bg=bg_c, fg=fg_c, center=True, bold=True)
            dcell(ws2, row, 4, iss.category,                    bg=bg_c, fg=fg_c)
            dcell(ws2, row, 5, iss.message,                     bg=bg_c, sz=10)
            dcell(ws2, row, 6, iss.location,                    bg=alt,  sz=9, italic=True, fg='666666')
            dcell(ws2, row, 7, iss.suggestion,                  bg='FEFEF0', sz=9, fg='444400')
            row += 1

    ws2.column_dimensions['A'].width = 7
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 10
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 50
    ws2.column_dimensions['F'].width = 22
    ws2.column_dimensions['G'].width = 45

    ws2.auto_filter.ref = 'A2:G2'
    ws2.freeze_panes   = 'C3'

    # ─────────────────────────────────────────────────────────────
    # Sheet 3: Thống kê theo nhóm lỗi
    # ─────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Thống kê lỗi")
    ws3.sheet_properties.tabColor = "2E6D45"

    ws3.merge_cells('A1:C1')
    c = ws3['A1']
    c.value = "THỐNG KÊ LỖI THEO NHÓM VÀ MỨC ĐỘ"
    c.font = XFont(name='Arial', size=12, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='1B4332')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 28

    for ci, h in enumerate(['Nhóm lỗi', 'Số lần xuất hiện', 'Mức độ phổ biến'], 1):
        hcell(ws3, 2, ci, h, bg='2D6A4F', sz=10)

    from collections import Counter
    cat_counts = Counter()
    for res in results:
        for iss in res.issues:
            if iss.severity in ("ERROR", "WARNING"):
                cat_counts[f"[{iss.severity}] {iss.category}"] += 1

    total_issues = sum(cat_counts.values()) or 1
    palette_green = ['D8F3DC', 'B7E4C7', '95D5B2', '74C69D', '52B788']

    for ri, (cat, cnt) in enumerate(cat_counts.most_common(), 1):
        ws3.row_dimensions[ri+2].height = 22
        bg = palette_green[ri % len(palette_green)]
        pct_str = f"{cnt/total_issues*100:.0f}%"
        dcell(ws3, ri+2, 1, cat,     bg=bg, bold=True)
        dcell(ws3, ri+2, 2, cnt,     bg=bg, center=True, bold=True, fg='1B4332')
        dcell(ws3, ri+2, 3, pct_str, bg=bg, center=True, fg='555555')

    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 18

    wb.save(output_path)
    print(f"\n✅ Đã lưu kết quả: {output_path}")


# ════════════════════════════════════════════════════════════════
#  IN BÁO CÁO RA CONSOLE (cho từng file)
# ════════════════════════════════════════════════════════════════
def print_report(result: CheckResult):
    filename = Path(result.filepath).name
    print(f"\n{'─'*70}")
    print(f"📄 File  : {filename}")
    if result.student_name:
        print(f"👤 SV    : {result.student_name}  |  MSSV: {result.student_id}")
    if result.advisor:
        print(f"🎓 GVHD  : {result.advisor}")
    if result.title:
        print(f"📝 Đề tài: {result.title[:80]}")
    print(f"📊 Điểm  : {result.score}/100  |  Lỗi: {result.error_count}  |  Cảnh báo: {result.warn_count}")
    print()

    sev_icons = {"ERROR": "❌", "WARNING": "⚠️ ", "INFO": "ℹ️ "}
    for iss in sorted(result.issues, key=lambda x: {"ERROR":0,"WARNING":1,"INFO":2}[x.severity]):
        icon = sev_icons.get(iss.severity, "  ")
        loc  = f" [{iss.location}]" if iss.location else ""
        print(f"  {icon} [{iss.category}]{loc}")
        print(f"      {iss.message}")
        if iss.suggestion:
            print(f"      → {iss.suggestion}")
    print()


# ════════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════════
def main():
    print("\n" + "="*70)
    print("  CHECK FORMAT KLTN — Kiểm tra định dạng Khóa Luận Tốt Nghiệp")
    print("  Trường Đại học Thủy Lợi · Khoa Kinh tế & QTKD")
    print("="*70)
    print("  Tiêu chuẩn kiểm tra:")
    print("  • Khổ giấy A4  • Lề: Trái 3cm, Phải 2cm, Trên/Dưới 2.5cm")
    print("  • Font Times New Roman  • Cỡ chữ theo quy định (13pt nội dung)")
    print("  • Giãn dòng 1.5  • Canh đều hai bên  • Cấu trúc bắt buộc")
    print("="*70 + "\n")

    # Lấy đường dẫn
    if len(sys.argv) > 1:
        target = sys.argv[1].strip('"').strip("'")
    else:
        try:
            import tkinter as tk
            from tkinter import filedialog, messagebox
            root_tk = tk.Tk()
            root_tk.withdraw()
            root_tk.title("Kiểm tra KLTN")
            choice = messagebox.askquestion(
                "Loại kiểm tra",
                "Chọn YES để quét toàn bộ THƯ MỤC\n"
                "Chọn NO để kiểm tra một FILE duy nhất"
            )
            if choice == 'yes':
                target = filedialog.askdirectory(title='Chọn thư mục chứa các file KLTN')
            else:
                target = filedialog.askopenfilename(
                    title='Chọn file KLTN (.docx)',
                    filetypes=[("Word Document", "*.docx"), ("All files", "*.*")]
                )
            root_tk.destroy()
            if not target:
                print("❌ Không chọn file/thư mục. Thoát.")
                sys.exit(0)
        except Exception:
            target = input("Nhập đường dẫn file hoặc thư mục: ").strip().strip('"')

    target_path = Path(target)
    if not target_path.exists():
        print(f"❌ Không tồn tại: {target}")
        sys.exit(1)

    # Quét
    if target_path.is_dir():
        print(f"📂 Thư mục: {target_path}")
        results = scan_directory(str(target_path))
        out_dir = target_path
    else:
        print(f"📄 File: {target_path.name}")
        results = [check_file(str(target_path))]
        out_dir = target_path.parent

    if not results:
        print("⚠️  Không tìm thấy file .docx nào!")
        sys.exit(0)

    # In báo cáo console
    for res in results:
        print_report(res)

    # Tóm tắt
    total  = len(results)
    passed = sum(1 for r in results if r.score >= 70)
    errors = sum(r.error_count for r in results)
    warns  = sum(r.warn_count  for r in results)

    print("="*70)
    print(f"📊 TỔNG KẾT: {total} file  |  Đạt (≥70đ): {passed}  |  Không đạt: {total-passed}")
    print(f"   Tổng lỗi: {errors}  |  Tổng cảnh báo: {warns}")
    print("="*70)

    # Xuất Excel
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = str(out_dir / f"KIEM_TRA_DINH_DANG_KLTN_{timestamp}.xlsx")
    export_excel(results, output_path)
    print(f"\n🎉 Hoàn tất! Kết quả lưu tại:\n   {output_path}\n")

    # Mở Excel (Mac/Windows)
    if sys.platform == 'darwin':
        try:
            os.system(f'open "{output_path}"')
        except Exception:
            pass
    elif sys.platform == 'win32':
        try:
            os.startfile(output_path)
        except Exception:
            pass


if __name__ == '__main__':
    main()
